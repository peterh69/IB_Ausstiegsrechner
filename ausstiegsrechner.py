"""
IB Ausstiegsrechner - Interactive Brokers Portfolio Analyzer
============================================================

Dieses Programm verbindet sich mit der Interactive Brokers Trader Workstation (TWS)
über die offizielle IB API und liest alle offenen Positionen sowie das Kontoguthaben
aus. Die Ergebnisse werden tabellarisch in einer GUI angezeigt und können optional
als Excel-Datei exportiert werden.

Unterstützte Positionstypen
---------------------------
- Cash Secured Puts (CSPs): Short-Put-Optionen mit Strike, DTE, Prämie und Restrendite
- Aktien (STK): Aktienpositionen mit aktuellem Kurs und Einstandspreis
- Covered Calls (CC): Short-Call-Optionen mit Strike, DTE, Prämie und Restrendite

Struktur der Anzeige
--------------------
1. Guthaben-Übersicht: EUR/USD Gesamtguthaben, für CSPs gebundenes Kapital, freier Cash
2. Cash Secured Puts: sortiert nach Symbol (alphabetisch) und DTE (aufsteigend)
3. Aktien & Covered Calls: getrennt nach EUR- und USD-Positionen, je alphabetisch

Restrendite-Berechnung
----------------------
Die annualisierte Restrendite wird nur berechnet, wenn eine Gewinnposition vorliegt
(aktueller Optionspreis < erhaltene Prämie):

    Restrendite p.a. = (365 / DTE) × (Erhaltene Prämie / Strike)

Freier Cash
-----------
Freier Cash = Gesamtguthaben (je Währung) − für CSPs gebundenes Kapital
Gebundenes Kapital je CSP = Strike × 100 × |Anzahl Kontrakte|

Voraussetzungen
---------------
1. Interactive Brokers TWS oder IB Gateway muss gestartet sein.
2. Die API-Verbindung muss in TWS aktiviert sein:
   File → Global Configuration → API → Settings → „Enable ActiveX and Socket Clients"
3. Python-Pakete: ib_insync, openpyxl (siehe requirements.txt)

Installation
------------
    pip install -r requirements.txt

Verwendung
----------
    python ausstiegsrechner.py

Konfiguration
-------------
Die Verbindungsparameter (Host, Port, Client-ID) können am Anfang der Datei
unter „Konfiguration" angepasst werden.

Ports:
    7496 = TWS Live-Trading
    7497 = TWS Paper-Trading
    4001 = IB Gateway Live
    4002 = IB Gateway Paper
"""

import asyncio
import logging
import threading
import tkinter as tk
from tkinter import ttk, messagebox
import time
from datetime import datetime, date

# Python 3.10+ erstellt keine Event Loop mehr automatisch.
# Vor dem Import von ib_insync muss eine neue Loop angelegt werden,
# damit asyncio-basierte Funktionen von ib_insync korrekt funktionieren.
asyncio.set_event_loop(asyncio.new_event_loop())

from ib_insync import IB, Forex, Stock, Option

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Log-Filter: Harmlose IB-Fehlercodes unterdrücken
# ---------------------------------------------------------------------------

# Diese Fehlercodes sind von IB als reine Informationsmeldungen gedacht
# und stellen keine echten Fehler dar. Sie werden aus dem Log herausgefiltert,
# um die Konsolenausgabe übersichtlich zu halten.
_BENIGN_LOG_CODES = {321, 354, 2104, 2106, 2107, 2108, 2158, 10090, 10167}


class _IBLogFilter(logging.Filter):
    """Unterdrückt bekannte, harmlose ib_insync-Fehlermeldungen im Log."""

    def filter(self, record):
        """Gibt False zurück wenn die Meldung einen harmlosen Fehlercode enthält."""
        msg = record.getMessage()
        return not any(
            f'Error {c},' in msg or f'{c},' in msg[:20]
            for c in _BENIGN_LOG_CODES
        )


# Filter auf alle relevanten ib_insync-Logger anwenden
for _log_name in ('ib_insync.wrapper', 'ib_insync.client', 'ib_insync.ib'):
    logging.getLogger(_log_name).addFilter(_IBLogFilter())


# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------

APP_VERSION = '0.15'       # Wird bei jeder Code-Änderung um 0.01 erhöht

TWS_HOST = '127.0.0.1'    # Hostname oder IP-Adresse der TWS/Gateway-Instanz
TWS_PORT = 7496            # API-Port (7496=TWS Live, 7497=TWS Paper, 4001=Gateway Live)
CLIENT_ID = 10             # Eindeutige Client-ID; darf nicht doppelt vergeben sein
OUTPUT_FILE = 'IB_Positionen.xlsx'  # Dateiname der Excel-Ausgabe
MARKET_DATA_WAIT = 3       # Sekunden, die auf eingehende Marktdaten gewartet wird


# ---------------------------------------------------------------------------
# Farbpalette für die Excel-Formatierung
# ---------------------------------------------------------------------------

COLOR_RED_HEADER   = 'C0504D'   # Dunkelrot  – Überschrift CSP-Abschnitt
COLOR_BLUE_HEADER  = '17375E'   # Dunkelblau – Überschrift Aktien/Calls-Abschnitt
COLOR_BLUE_STOCK   = 'DCE6F1'   # Hellblau   – Hintergrund Aktienzeile
COLOR_GREEN_CALL   = 'EBF1DE'   # Hellgrün   – Hintergrund Call-Zeile
COLOR_SEPARATOR    = 'D9D9D9'   # Grau       – Trennzeile zwischen Symbolen
COLOR_CSP_ROW_1    = 'FFFFFF'   # Weiß       – CSP-Zeile (gerader Index)
COLOR_CSP_ROW_2    = 'F2DCDB'   # Rosa       – CSP-Zeile (ungerader Index, alternierend)
COLOR_CASH_HEADER  = '375623'   # Dunkelgrün – Überschrift Guthaben-Abschnitt
COLOR_CASH_ROW     = 'EBF1DE'   # Hellgrün   – Guthaben-Datenzeilen
COLOR_EUR_GROUP    = 'FFF2CC'   # Hellgelb   – Unterabschnitts-Header EUR-Positionen
COLOR_USD_GROUP    = 'DEEAF1'   # Hellblau   – Unterabschnitts-Header USD-Positionen


# ---------------------------------------------------------------------------
# Hilfsfunktionen – Preisermittlung und Datumsformatierung
# ---------------------------------------------------------------------------

def get_price(ticker) -> float | None:
    """Ermittelt den besten verfügbaren Marktpreis aus einem ib_insync-Ticker.

    Prioritätsreihenfolge: last → close → Mitte aus Bid/Ask.
    Gibt None zurück, wenn kein Preis verfügbar ist.

    Args:
        ticker: ib_insync Ticker-Objekt oder None.

    Returns:
        Preis als float, oder None wenn kein Preis verfügbar.
    """
    if ticker is None:
        return None
    last  = ticker.last
    close = ticker.close
    bid   = ticker.bid
    ask   = ticker.ask

    if last and last > 0:
        return last
    if close and close > 0:
        return close
    if bid and ask and bid > 0 and ask > 0:
        return (bid + ask) / 2
    return None


def dte(expiry_str: str) -> int:
    """Berechnet die verbleibenden Tage bis zum Verfall (Days to Expiration).

    Args:
        expiry_str: Verfallsdatum im Format YYYYMMDD.

    Returns:
        Anzahl der verbleibenden Tage als int, oder 0 bei ungültigem Datum.
    """
    try:
        exp_date = datetime.strptime(expiry_str, '%Y%m%d').date()
        return (exp_date - date.today()).days
    except Exception:
        return 0


_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
           'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def fmt_option_symbol(symbol: str, expiry_str: str, strike: float, right: str) -> str:
    """Erzeugt eine lesbare Optionsbezeichnung im TWS-Stil.

    Format: <Symbol> <MonTT'JJ> <Strike> <Put|Call>
    Beispiel: ALV Mar13'26 370 Put

    Monatsnamen werden immer auf Englisch ausgegeben (unabhängig von der Systemsprache).

    Args:
        symbol: Ticker-Symbol (z.B. 'ALV').
        expiry_str: Verfallsdatum im Format YYYYMMDD.
        strike: Strike-Preis als float.
        right: Optionstyp – 'P' für Put, 'C' für Call.

    Returns:
        Lesbares Optionssymbol als String.
    """
    try:
        dt = datetime.strptime(expiry_str, '%Y%m%d')
        month = _MONTHS[dt.month - 1]
        day   = str(dt.day)           # kein führendes Null
        year  = dt.strftime('%y')     # 2-stelliges Jahr
        strike_str = f"{strike:g}"    # entfernt überflüssige Nullen (370.0 → '370')
        type_str   = 'Put' if right == 'P' else 'Call'
        return f"{symbol} {month}{day}'{year} {strike_str} {type_str}"
    except Exception:
        return symbol


def fmt_date(expiry_str: str) -> str:
    """Konvertiert ein Datum von YYYYMMDD nach DD.MM.YYYY (deutsches Format).

    Args:
        expiry_str: Datum als String im Format YYYYMMDD.

    Returns:
        Datum als String im Format DD.MM.YYYY, oder den Originalwert bei Fehler.
    """
    try:
        return datetime.strptime(expiry_str, '%Y%m%d').strftime('%d.%m.%Y')
    except Exception:
        return expiry_str


# ---------------------------------------------------------------------------
# Hilfsfunktionen – Excel-Formatierung
# ---------------------------------------------------------------------------

def apply_fill(cell, hex_color: str):
    """Setzt die Hintergrundfarbe einer Excel-Zelle.

    Args:
        cell: openpyxl-Cell-Objekt.
        hex_color: Farbe als 6-stelliger Hex-String (ohne führendes #).
    """
    cell.fill = PatternFill(fill_type='solid', fgColor=hex_color)


def apply_header_style(cell, hex_color: str):
    """Formatiert eine Zelle als Abschnitts-Header (farbiger Hintergrund, weiße Fettschrift).

    Args:
        cell: openpyxl-Cell-Objekt.
        hex_color: Hintergrundfarbe als 6-stelliger Hex-String.
    """
    cell.fill = PatternFill(fill_type='solid', fgColor=hex_color)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')


def apply_subgroup_style(cell, hex_color: str):
    """Formatiert eine Zelle als Untergruppen-Header (farbiger Hintergrund, schwarze Fettschrift).

    Wird für die EUR/USD-Gruppen-Unterüberschriften in Abschnitt 2 verwendet.

    Args:
        cell: openpyxl-Cell-Objekt.
        hex_color: Hintergrundfarbe als 6-stelliger Hex-String.
    """
    cell.fill = PatternFill(fill_type='solid', fgColor=hex_color)
    cell.font = Font(bold=True, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')


def thin_border() -> Border:
    """Erstellt einen dünnen grauen Rahmen für Trennzeilen.

    Returns:
        openpyxl Border-Objekt mit dünner unterer Linie.
    """
    thin = Side(style='thin', color='AAAAAA')
    return Border(bottom=thin)


# ---------------------------------------------------------------------------
# Hilfsfunktionen – Finanzberechnungen
# ---------------------------------------------------------------------------

def calc_restrendite(premium: float, strike: float, days: int,
                     current_price) -> float | None:
    """Berechnet die annualisierte Restrendite einer Short-Option-Position.

    Die Restrendite wird nur berechnet, wenn:
    - Ein gültiger Strike, eine positive Prämie und verbleibende Laufzeit vorliegen.
    - Der aktuelle Optionspreis unter der erhaltenen Prämie liegt (Gewinnposition).

    Formel:
        Restrendite p.a. = (365 / DTE) × (Erhaltene Prämie / Strike)

    Args:
        premium: Erhaltene Prämie pro Aktie (Einstiegsverkaufspreis).
        strike: Strike-Preis der Option.
        days: Verbleibende Tage bis zum Verfall (DTE).
        current_price: Aktueller Marktpreis der Option (zum Rückkauf), oder None.

    Returns:
        Restrendite als Dezimalzahl (z.B. 0.15 = 15% p.a.), oder None wenn
        keine Gewinnposition vorliegt oder die Berechnung nicht möglich ist.
    """
    if days <= 0 or strike <= 0 or premium <= 0:
        return None
    # Kein Gewinn: aktueller Preis >= erhaltene Prämie (Rückkauf wäre teurer)
    if current_price is not None and current_price >= premium:
        return None
    return (365.0 / days) * (premium / strike)


def fetch_long_names(ib: IB, stock_contracts: dict) -> dict:
    """Lädt die offiziellen Langnamen (Firmennamen) für alle übergebenen Kontrakte.

    Verwendet IB's reqContractDetails() um den vollständigen Firmennamen
    (z.B. 'NVIDIA Corporation') aus den Vertragsdaten zu lesen.

    Args:
        ib: Aktive IB-Verbindung.
        stock_contracts: Dict {symbol: Stock-Kontrakt}.

    Returns:
        Dict {symbol: long_name}. Bei Fehlern wird das Ticker-Symbol als Fallback verwendet.
    """
    long_names = {}
    for sym, contract in stock_contracts.items():
        try:
            details = ib.reqContractDetails(contract)
            if details and details[0].longName:
                long_names[sym] = details[0].longName
            else:
                long_names[sym] = sym
        except Exception:
            long_names[sym] = sym
    return long_names


def fetch_account_cash(ib: IB) -> dict:
    """Liest die Barguthaben aus dem IB-Konto, getrennt nach Währung.

    Liest den 'CashBalance'-Tag aus den Kontowerten für EUR und USD.
    Der CashBalance gibt das tatsächliche Barguthaben in der jeweiligen Währung an,
    unabhängig von Marginanforderungen.

    Args:
        ib: Aktive IB-Verbindung.

    Returns:
        Dict {'EUR': float, 'USD': float} mit den Barguthaben.
        Nicht vorhandene Währungen werden mit 0.0 vorbelegt.
    """
    # Alle Konten summieren (bei Multi-Account-Setup gibt es mehrere Einträge
    # pro Währung – der letzte Eintrag wäre sonst der einzig gezählte)
    cash = {'EUR': 0.0, 'USD': 0.0}
    for av in ib.accountValues():
        if av.tag == 'CashBalance' and av.currency in cash:
            try:
                cash[av.currency] += float(av.value)
            except ValueError:
                pass
    return cash


# ---------------------------------------------------------------------------
# Daten sammeln (IB-Verbindung erforderlich)
# ---------------------------------------------------------------------------

def collect_data(ib: IB, status_callback=None) -> dict:
    """Liest alle Positionen und Kontodaten aus IB und gibt sie als Dict zurück.

    Args:
        ib: Aktive, bereits verbundene IB-Instanz.
        status_callback: Optionale Funktion(str) für Statusmeldungen.

    Returns:
        Dict mit allen Positionsdaten (cash_balance, csp_rows, stock_map, etc.)
        oder None wenn keine Positionen gefunden wurden.
    """
    def status(msg: str):
        if status_callback:
            status_callback(msg)

    # --- Kontoguthaben abrufen ---
    status('Lese Kontoguthaben...')
    cash_balance = fetch_account_cash(ib)

    # --- EUR/USD-Wechselkurs anfordern ---
    eurusd_contract = Forex('EURUSD')
    ib.qualifyContracts(eurusd_contract)
    eurusd_ticker = ib.reqMktData(eurusd_contract, '', False, False)

    # --- Positionen laden ---
    positions = ib.positions()
    if not positions:
        return None

    # --- Positionen klassifizieren ---
    csps   = []   # Short Puts (Cash Secured Puts)
    stocks = []   # Aktienpositionen
    calls  = []   # Short Calls (Covered Calls)

    for pos in positions:
        contract = pos.contract
        sec_type = contract.secType
        right = getattr(contract, 'right', '')

        if sec_type == 'OPT' and right == 'P':
            csps.append(pos)
        elif sec_type == 'OPT' and right == 'C':
            calls.append(pos)
        elif sec_type == 'STK':
            stocks.append(pos)

    # --- Underlying-Kontrakte für Optionen erzeugen ---
    underlying_symbols = set()
    for pos in csps + calls:
        c = pos.contract
        underlying_symbols.add((c.symbol, getattr(c, 'currency', 'USD')))

    underlying_contracts = {}
    for sym, cur in underlying_symbols:
        stk = Stock(sym, 'SMART', cur)
        try:
            ib.qualifyContracts(stk)
        except Exception:
            pass
        underlying_contracts[(sym, cur)] = stk

    # --- Marktdaten gebündelt anfordern ---
    status('Frage Marktdaten ab...')

    # Optionen: qualifizieren und Marktdaten abonnieren
    opt_tickers = {}
    for pos in csps + calls:
        c = pos.contract
        try:
            ib.qualifyContracts(c)
        except Exception:
            pass
        t = ib.reqMktData(c, '', False, False)
        opt_tickers[c.conId] = t

    # Aktien: qualifizieren und Marktdaten abonnieren
    stk_tickers = {}
    for pos in stocks:
        c = pos.contract
        try:
            ib.qualifyContracts(c)
        except Exception:
            pass
        t = ib.reqMktData(c, '', False, False)
        stk_tickers[c.conId] = t

    # Underlying-Kurse für Optionen abonnieren
    underlying_tickers = {}
    for key, stk in underlying_contracts.items():
        t = ib.reqMktData(stk, '', False, False)
        underlying_tickers[key] = t

    # Auf Marktdaten warten
    status(f'Warte {MARKET_DATA_WAIT}s auf Marktdaten...')
    time.sleep(MARKET_DATA_WAIT)
    ib.sleep(0)  # ib_insync-Event-Loop einen Verarbeitungszyklus ausführen lassen

    # --- Firmennamen laden ---
    status('Lade Firmennamen...')
    all_known_stocks = {}
    for (sym, _cur), contract in underlying_contracts.items():
        all_known_stocks[sym] = contract
    for pos in stocks:
        sym = pos.contract.symbol
        if sym not in all_known_stocks:
            all_known_stocks[sym] = pos.contract

    long_names = fetch_long_names(ib, all_known_stocks)

    # --- EUR/USD-Kurs auslesen ---
    eurusd_price = get_price(eurusd_ticker)
    if eurusd_price is None:
        eurusd_price = 1.0

    # --- CSP-Daten sammeln ---
    csp_rows = []
    for pos in csps:
        c = pos.contract
        sym = c.symbol
        cur = getattr(c, 'currency', 'USD')

        premium_per_share = abs(pos.avgCost) / 100.0

        opt_ticker        = opt_tickers.get(c.conId)
        current_opt_price = get_price(opt_ticker)

        ul_ticker        = underlying_tickers.get((sym, cur))
        underlying_price = get_price(ul_ticker)

        strike = float(getattr(c, 'strike', 0))
        expiry = getattr(c, 'lastTradeDateOrContractMonth', '')
        days   = dte(expiry)

        csp_rows.append({
            'symbol':           sym,
            'display_symbol':   fmt_option_symbol(sym, expiry, strike, 'P'),
            'bezeichnung':      long_names.get(sym, sym),
            'position':         pos.position,
            'underlying_price': underlying_price,
            'strike':           strike,
            'dte':              days,
            'expiry':           expiry,
            'premium':          premium_per_share,
            'current_price':    current_opt_price,
            'eurusd':           eurusd_price,
            'currency':         cur,
        })

    # Sortierung: erst USD, dann EUR – innerhalb jeder Währung alphabetisch nach Symbol, dann DTE
    csp_rows.sort(key=lambda r: (0 if r['currency'] == 'USD' else 1, r['symbol'], r['dte']))

    # --- In CSPs gebundenes Kapital ---
    csp_margin = {'EUR': 0.0, 'USD': 0.0}
    for row in csp_rows:
        cur = row['currency']
        if cur in csp_margin:
            csp_margin[cur] += row['strike'] * 100 * abs(row['position'])

    # --- Aktien-Daten sammeln ---
    stock_map = {}
    for pos in stocks:
        c = pos.contract
        sym = c.symbol
        cur = getattr(c, 'currency', 'USD')

        # Preis aus stk_tickers (pos.contract, z.B. Exchange IBIS) holen.
        # Fallback: underlying_tickers (SMART-gerouteter Kontrakt), der für
        # Optionen bereits erfolgreich Preise liefert.
        current_price = get_price(stk_tickers.get(c.conId))
        if current_price is None:
            current_price = get_price(underlying_tickers.get((sym, cur)))

        if sym in stock_map:
            # Mehrere IB-Einträge für dasselbe Symbol (z.B. verschiedene Exchanges
            # oder Teilpositionen) → Stückzahl addieren; Preis und avgCost vom
            # ersten Eintrag mit gültigem Kurs behalten.
            stock_map[sym]['position'] += pos.position
            if stock_map[sym]['current_price'] is None and current_price is not None:
                stock_map[sym]['current_price'] = current_price
        else:
            stock_map[sym] = {
                'symbol':         sym,
                'display_symbol': sym,
                'bezeichnung':    long_names.get(sym, sym),
                'position':       pos.position,
                'avg_cost':       pos.avgCost,
                'current_price':  current_price,
                'eurusd':         eurusd_price,
                'currency':       cur,
            }

    # --- Call-Daten sammeln ---
    call_map = {}
    for pos in calls:
        c = pos.contract
        sym = c.symbol
        cur = getattr(c, 'currency', 'USD')

        premium_per_share = abs(pos.avgCost) / 100.0

        opt_ticker        = opt_tickers.get(c.conId)
        current_opt_price = get_price(opt_ticker)

        strike = float(getattr(c, 'strike', 0))
        expiry = getattr(c, 'lastTradeDateOrContractMonth', '')
        days   = dte(expiry)

        is_long = pos.position > 0  # True = gekaufter Call, False = Covered Call (Short)

        row = {
            'symbol':         sym,
            'display_symbol': fmt_option_symbol(sym, expiry, strike, 'C'),
            'bezeichnung':    long_names.get(sym, sym),
            'position':       pos.position,
            'is_long':        is_long,
            'strike':         strike,
            'dte':            days,
            'expiry':         expiry,
            'premium':        premium_per_share,
            'current_price':  current_opt_price,
            'eurusd':         eurusd_price,
            'currency':       cur,
        }
        call_map.setdefault(sym, []).append(row)

    # Calls innerhalb eines Symbols nach DTE aufsteigend sortieren
    for sym in call_map:
        call_map[sym].sort(key=lambda r: r['dte'])

    # --- In Aktien gebundenes Kapital (Einstandswert) ---
    stock_capital = {'EUR': 0.0, 'USD': 0.0}
    for s in stock_map.values():
        cur = s['currency']
        if cur in stock_capital:
            stock_capital[cur] += s['position'] * s['avg_cost']

    # --- Fehlende Underlying-Preise für CSPs/Calls aus stock_map ergänzen ---
    # Wenn SMART-Routing keinen Preis liefert, aber die Aktie im Depot liegt,
    # deren Kurs als Fallback nutzen (z.B. RGLD: SMART-Ticker ohne Daten,
    # aber Aktienposition mit gültigem Kurs vorhanden).
    for row in csp_rows:
        if row['underlying_price'] is None and row['symbol'] in stock_map:
            row['underlying_price'] = stock_map[row['symbol']]['current_price']

    # --- In Long Calls gebundenes Kapital (bezahlte Prämie × 100 × Kontrakte) ---
    long_call_capital = {'EUR': 0.0, 'USD': 0.0}
    for call_list in call_map.values():
        for row in call_list:
            if row.get('is_long'):
                cur = row['currency']
                if cur in long_call_capital:
                    long_call_capital[cur] += row['premium'] * 100 * abs(row['position'])

    # --- Freier Cash = Barmittel − CSP-Kapital ---
    # (Aktien- und Long Call-Kapital werden nur informativ angezeigt)
    free_cash = {
        cur: cash_balance.get(cur, 0.0) - csp_margin.get(cur, 0.0)
        for cur in ('EUR', 'USD')
    }

    # --- Symbole nach Währung aufteilen ---
    all_syms_2 = sorted(set(list(stock_map.keys()) + list(call_map.keys())))

    def sym_currency(sym: str) -> str:
        if sym in stock_map:
            return stock_map[sym]['currency']
        if sym in call_map and call_map[sym]:
            return call_map[sym][0]['currency']
        return 'USD'

    syms_eur = [sym for sym in all_syms_2 if sym_currency(sym) == 'EUR']
    syms_usd = [sym for sym in all_syms_2 if sym_currency(sym) != 'EUR']

    return {
        'cash_balance':      cash_balance,
        'csp_margin':        csp_margin,
        'stock_capital':     stock_capital,
        'long_call_capital': long_call_capital,
        'free_cash':         free_cash,
        'csp_rows':          csp_rows,
        'stock_map':    stock_map,
        'call_map':     call_map,
        'syms_eur':     syms_eur,
        'syms_usd':     syms_usd,
        'eurusd_price': eurusd_price,
        'timestamp':    datetime.now().strftime('%d.%m.%Y %H:%M'),
    }


# ---------------------------------------------------------------------------
# Excel-Export
# ---------------------------------------------------------------------------

def write_excel(data: dict, filename: str):
    """Erstellt eine Excel-Datei aus den gesammelten Positionsdaten.

    Args:
        data: Dict wie von collect_data() zurückgegeben.
        filename: Pfad/Dateiname der zu schreibenden Excel-Datei.
    """
    cash_balance = data['cash_balance']
    csp_margin   = data['csp_margin']
    free_cash    = data['free_cash']
    csp_rows     = data['csp_rows']
    stock_map    = data['stock_map']
    call_map     = data['call_map']
    syms_eur     = data['syms_eur']
    syms_usd     = data['syms_usd']
    eurusd_price = data['eurusd_price']
    now_str      = data['timestamp']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Positionen'

    current_row = 1
    NUM_COLS = 12

    # --- Titelzeile ---
    title_text = f"IB Positionen  |  Stand: {now_str}  |  EUR/USD: {eurusd_price:.4f}"
    title_cell = ws.cell(row=current_row, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=NUM_COLS)
    current_row += 2

    # ===========================================================
    # ABSCHNITT 0: Guthaben-Übersicht
    # ===========================================================

    ws.cell(row=current_row, column=1, value='Kontoguthaben')
    for col in range(1, NUM_COLS + 1):
        apply_header_style(ws.cell(row=current_row, column=col), COLOR_CASH_HEADER)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=NUM_COLS)
    current_row += 1

    for col_idx, header in enumerate(['', 'EUR', 'USD'], start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        apply_fill(cell, 'C6EFCE')
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    cash_rows_data = [
        ('Gesamtguthaben', cash_balance['EUR'], cash_balance['USD']),
        ('CSP-Kapital',    csp_margin['EUR'],   csp_margin['USD']),
        ('Freier Cash',    free_cash['EUR'],     free_cash['USD']),
    ]
    for label, eur_val, usd_val in cash_rows_data:
        is_free_cash = label == 'Freier Cash'
        cell_label = ws.cell(row=current_row, column=1, value=label)
        cell_label.font = Font(bold=is_free_cash)
        apply_fill(cell_label, COLOR_CASH_ROW)

        for col_idx, val in enumerate([eur_val, usd_val], start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(bold=is_free_cash,
                             color='FF0000' if is_free_cash and val < 0 else '000000')
            apply_fill(cell, COLOR_CASH_ROW)
        current_row += 1

    current_row += 1  # Leerzeile

    # ===========================================================
    # ABSCHNITT 1: Cash Secured Puts (CSPs)
    # ===========================================================

    csp_headers = [
        'Symbol', 'Bezeichnung', 'Position', 'Kurs Underlying',
        'Strike', 'DTE', 'Ablaufdatum',
        'Erhaltene Prämie', 'Akt. Options-Preis', 'Restrendite p.a.', 'Währung'
    ]
    NUM_CSP_COLS = len(csp_headers)

    ws.cell(row=current_row, column=1, value='Cash Secured Puts (CSPs)')
    for col in range(1, NUM_CSP_COLS + 1):
        apply_header_style(ws.cell(row=current_row, column=col), COLOR_RED_HEADER)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=NUM_CSP_COLS)
    current_row += 1

    for col_idx, header in enumerate(csp_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='F2DCDB')
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for i, row in enumerate(csp_rows):
        bg = COLOR_CSP_ROW_1 if i % 2 == 0 else COLOR_CSP_ROW_2
        ul_price    = row['underlying_price']
        cur_opt     = row['current_price']
        restrendite = calc_restrendite(row['premium'], row['strike'],
                                       row['dte'], cur_opt)

        values = [
            row['display_symbol'],
            row['bezeichnung'],
            row['position'],
            ul_price if ul_price is not None else 'n/v',
            row['strike'],
            row['dte'],
            fmt_date(row['expiry']),
            row['premium'],
            cur_opt if cur_opt is not None else 'n/v',
            restrendite if restrendite is not None else '-',
            row['currency'],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            apply_fill(cell, bg)
            cell.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left')
            if col_idx in (4, 5, 8, 9) and isinstance(val, float):
                cell.number_format = '#,##0.00'
            elif col_idx == 10 and isinstance(val, float):
                cell.number_format = '0.00%'
        current_row += 1

    current_row += 1  # Leerzeile

    # ===========================================================
    # ABSCHNITT 2: Aktien & Covered Calls
    # ===========================================================

    sec2_headers = [
        'Symbol', 'Bezeichnung', 'Typ', 'Position', 'Akt. Kurs',
        'Strike', 'DTE', 'Ablaufdatum',
        'Kauf-/Verkaufspreis', 'Akt. Options-Preis', 'Restrendite p.a.', 'Währung'
    ]
    NUM_SEC2_COLS = len(sec2_headers)

    ws.cell(row=current_row, column=1, value='Aktien & Covered Calls')
    for col in range(1, NUM_SEC2_COLS + 1):
        apply_header_style(ws.cell(row=current_row, column=col), COLOR_BLUE_HEADER)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=NUM_SEC2_COLS)
    current_row += 1

    for col_idx, header in enumerate(sec2_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='DCE6F1')
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    def write_sym_group(sym_list: list, group_label: str, group_color: str):
        nonlocal current_row

        if not sym_list:
            return

        ws.cell(row=current_row, column=1, value=group_label)
        for col in range(1, NUM_SEC2_COLS + 1):
            apply_subgroup_style(ws.cell(row=current_row, column=col), group_color)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=NUM_SEC2_COLS)
        current_row += 1

        for sym_idx, sym in enumerate(sym_list):

            if sym in stock_map:
                s = stock_map[sym]
                cur_price = s['current_price']
                values = [
                    s['display_symbol'],
                    s['bezeichnung'],
                    'Aktie',
                    s['position'],
                    cur_price if cur_price is not None else 'n/v',
                    '-', '-', '-',
                    s['avg_cost'],
                    '-', '-',
                    s['currency'],
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    apply_fill(cell, COLOR_BLUE_STOCK)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left')
                    if col_idx == 5 and isinstance(val, float):
                        cell.number_format = '#,##0.00'
                    if col_idx == 9 and isinstance(val, float):
                        cell.number_format = '#,##0.00'
                current_row += 1

            if sym in call_map:
                for call_row in call_map[sym]:
                    cur_opt     = call_row['current_price']
                    restrendite = calc_restrendite(
                        call_row['premium'], call_row['strike'],
                        call_row['dte'], cur_opt
                    )
                    values = [
                        call_row['display_symbol'],
                        call_row['bezeichnung'],
                        'Call',
                        call_row['position'],
                        '-',
                        call_row['strike'],
                        call_row['dte'],
                        fmt_date(call_row['expiry']),
                        call_row['premium'],
                        cur_opt if cur_opt is not None else 'n/v',
                        restrendite if restrendite is not None else '-',
                        call_row['currency'],
                    ]
                    for col_idx, val in enumerate(values, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        apply_fill(cell, COLOR_GREEN_CALL)
                        cell.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left')
                        if col_idx in (6, 9, 10) and isinstance(val, float):
                            cell.number_format = '#,##0.00'
                        if col_idx == 11 and isinstance(val, float):
                            cell.number_format = '0.00%'
                    current_row += 1

            if sym_idx < len(sym_list) - 1:
                for col in range(1, NUM_SEC2_COLS + 1):
                    cell = ws.cell(row=current_row, column=col)
                    apply_fill(cell, COLOR_SEPARATOR)
                    cell.border = thin_border()
                current_row += 1

        current_row += 1  # Leerzeile nach der Gruppe

    write_sym_group(syms_eur, 'EUR-Positionen', COLOR_EUR_GROUP)
    write_sym_group(syms_usd, 'USD-Positionen', COLOR_USD_GROUP)

    # --- Spaltenbreiten automatisch anpassen ---
    for col in range(1, NUM_COLS + 1):
        max_width = 10
        col_letter = get_column_letter(col)
        for row_cells in ws.iter_rows(min_col=col, max_col=col):
            for cell in row_cells:
                if cell.value and cell.data_type != 'n':
                    try:
                        cell_len = len(str(cell.value))
                        if cell_len > max_width:
                            max_width = cell_len
                    except Exception:
                        pass
        ws.column_dimensions[col_letter].width = min(max_width + 2, 40)

    wb.save(filename)


# ---------------------------------------------------------------------------
# CSP-Kandidaten suchen (für CSP-Auswahl-Dialog)
# ---------------------------------------------------------------------------

# Client-ID für die zweite IB-Verbindung im CSP-Auswahl-Dialog
CSP_CLIENT_ID = CLIENT_ID + 1

# Maximale Anzahl von Strikes pro Verfallstermin (begrenzt Marktdaten-Abfragen)
CSP_MAX_STRIKES_PER_EXPIRY = 8

# Kursbereich für Strikes relativ zum aktuellen Kurs: von 70% bis 102%
CSP_STRIKE_MIN_FACTOR = 0.70
CSP_STRIKE_MAX_FACTOR = 1.02


def fetch_csp_candidates(ib: IB, ticker: str, loaded_data: dict,
                         status_callback=None) -> dict:
    """Lädt verfügbare Put-Optionen (CSP-Kandidaten) für einen Ticker.

    Sucht Short-Put-Optionen für die nächsten 8 Wochen (je einen Verfallstermin
    pro Woche, nächster Freitag ± 4 Tage) und Strikes im Bereich 70–102% des
    aktuellen Kurses. Berechnet Restrendite und %-Abstand zum Kurs.

    Restrendite  = (365 / DTE) × (Bid / Strike)
    % zum Kurs   = (Strike / aktueller Kurs − 1) × 100

    Ablauf:
        1. Underlying-Kontrakt bestimmen (aus geladenem Portfolio oder neu qualifizieren)
        2. Aktuellen Kurs und Firmennamen laden
        3. Optionskette via reqSecDefOptParams abrufen
        4. Nächste 8 Freitage bestimmen, je nächstliegenden Verfallstermin wählen
        5. Strikes filtern (70–102% des Kurses), Put-Kontrakte qualifizieren
        6. Marktdaten abrufen, Ergebnisse nach Woche / Strike sortieren

    Args:
        ib: Aktive, bereits verbundene IB-Instanz.
        ticker: Ticker-Symbol (z.B. 'AAPL', 'DTE').
        loaded_data: Dict wie von collect_data() zurückgegeben; wird für
                     Exchange-/Währungs-Hints genutzt. Kann None sein.
        status_callback: Optionale Funktion(str) für Statusmeldungen.

    Returns:
        Dict mit:
            'ticker':        Ticker-Symbol
            'long_name':     Firmenname
            'current_price': Aktueller Kurs des Underlyings
            'currency':      Währung (EUR/USD)
            'options':       Liste von Dicts je Option:
                               symbol, strike, pct_to_price, dte, expiry,
                               bid, restrendite, exchange
        Nur Optionen mit gültigem Bid-Preis sind enthalten.

    Raises:
        ValueError: Wenn Ticker nicht gefunden, kein Kurs oder keine Optionen verfügbar.
    """
    def status(msg: str):
        if status_callback:
            status_callback(msg)

    # --- Underlying-Kontrakt bestimmen ---
    # Priorität:
    #   1. Geladenes Portfolio – stock_map (direkte Aktienposition)
    #   2. Geladenes Portfolio – csp_rows / call_map (Währungs-Hint aus Optionspositionen)
    #   3. IB-Positionen (STK) im verbundenen Konto ("Peter Sammlung")
    #   4. Fallback: SMART/EUR, dann SMART/USD
    stock_contract = None
    currency = 'USD'

    # 1. Direkte Aktienposition im geladenen Portfolio
    if loaded_data and ticker in loaded_data.get('stock_map', {}):
        s = loaded_data['stock_map'][ticker]
        currency = s.get('currency', 'USD')
        stock_contract = Stock(ticker, 'SMART', currency)

    # 2. Währungs-Hint aus CSP- oder Call-Positionen (Ticker als Underlying)
    if stock_contract is None and loaded_data:
        for row in loaded_data.get('csp_rows', []):
            if row['symbol'] == ticker:
                currency = row.get('currency', 'USD')
                stock_contract = Stock(ticker, 'SMART', currency)
                break
        if stock_contract is None:
            for sym, calls in loaded_data.get('call_map', {}).items():
                if sym == ticker and calls:
                    currency = calls[0].get('currency', 'USD')
                    stock_contract = Stock(ticker, 'SMART', currency)
                    break

    # 3. Alle IB-Positionen prüfen (STK-Positionen im verbundenen Konto)
    if stock_contract is None:
        for pos in ib.positions():
            c = pos.contract
            if c.symbol == ticker and c.secType == 'STK':
                stock_contract = c
                currency = getattr(c, 'currency', 'USD')
                break

    # 4. Fallback: SMART/EUR zuerst (häufiger für deutsche Aktien), dann SMART/USD
    if stock_contract is None:
        for cur in ('EUR', 'USD'):
            candidate = Stock(ticker, 'SMART', cur)
            qualified = ib.qualifyContracts(candidate)
            if qualified:
                stock_contract = qualified[0]
                currency = cur
                break

    if stock_contract is None:
        raise ValueError(f'Ticker "{ticker}" bei Interactive Brokers nicht gefunden.')

    # Kontrakt qualifizieren (conId ermitteln, wird für reqSecDefOptParams benötigt)
    qualified = ib.qualifyContracts(stock_contract)
    if not qualified:
        raise ValueError(f'Kontrakt für "{ticker}" konnte nicht qualifiziert werden.')
    stock_contract = qualified[0]

    # --- Firmennamen und aktuellen Kurs laden ---
    status(f'{ticker}: Lade Kurs und Firmenname...')
    long_name = fetch_long_names(ib, {ticker: stock_contract}).get(ticker, ticker)
    stk_ticker_obj = ib.reqMktData(stock_contract, '', False, False)
    time.sleep(2)
    ib.sleep(0)
    current_price = get_price(stk_ticker_obj)
    # Börsenschluss erkennen: kein live Bid/Ask, aber Last-Preis vorhanden
    stk_bid = stk_ticker_obj.bid
    stk_ask = stk_ticker_obj.ask
    is_market_closed = (
        current_price is not None and
        (not stk_bid or stk_bid <= 0) and
        (not stk_ask or stk_ask <= 0)
    )
    ib.cancelMktData(stock_contract)

    if current_price is None:
        raise ValueError(
            f'Kein aktueller Kurs für "{ticker}" verfügbar.\n'
            'Ist der Markt geöffnet oder gibt es Delayed-Daten?'
        )

    status(f'{ticker}: Kurs {current_price:.2f} {currency}. Lade Optionskette...')

    # --- Optionskette laden ---
    chains = ib.reqSecDefOptParams(ticker, '', 'STK', stock_contract.conId)

    if not chains:
        raise ValueError(f'Keine Optionskette für "{ticker}" verfügbar.')

    # --- Besten Exchange wählen und ALLE zugehörigen Chains einsammeln ---
    # IB liefert für einen Underlying oft mehrere OptionChain-Objekte mit demselben Exchange
    # aber unterschiedlicher tradingClass – z.B. für EUREX/DTB:
    #   'DTE'  = monatliche Optionen (3. Freitag des Monats)
    #   'DTE1' = wöchentliche Optionen (jeden Freitag, Woche 1)
    #   'DTE2' = wöchentliche Optionen (Woche 2), usw.
    # Nur durch Zusammenführen aller Chains erhält man alle Verfallstermine.
    preferred_exchanges = ['CBOE', 'DTB', 'EUREX', 'BOX', 'SMART']
    best_exchange = None
    for pref in preferred_exchanges:
        if any(c.exchange == pref for c in chains):
            best_exchange = pref
            break
    if best_exchange is None:
        best_exchange = chains[0].exchange

    # Alle Chains dieses Exchange (monatlich + wöchentlich)
    exchange_chains = [c for c in chains if c.exchange == best_exchange]

    # Expirations und Strikes aus ALLEN Chains zusammenführen.
    # expiry_to_chain: merkt sich welche Chain (tradingClass/multiplier) zu welchem
    # Verfallstermin gehört – nötig für korrekten Option-Kontrakt später.
    expiry_to_chain: dict = {}
    all_strikes_set: set = set()
    for c in exchange_chains:
        for exp in c.expirations:
            if exp not in expiry_to_chain:   # erste (= monatliche) Chain hat Vorrang
                expiry_to_chain[exp] = c
        all_strikes_set.update(c.strikes)

    status(
        f'{ticker}: Exchange {best_exchange}, '
        f'{len(exchange_chains)} Serien, '
        f'{len(expiry_to_chain)} Verfallstermine gesamt. '
        f'Wähle nächste 8...'
    )

    # --- Nächste 8 Verfallstermine (chronologisch) ---
    # Einfach die 8 frühesten verfügbaren Expirations nehmen – das ergibt automatisch
    # wöchentliche Abstände wenn wöchentliche Optionen vorhanden sind, und monatliche
    # wenn nur Monatsoptionen verfügbar sind.
    all_available = sorted(
        [exp for exp in expiry_to_chain if dte(exp) > 0],
        key=lambda e: dte(e)
    )
    valid_expirations = all_available[:8]

    if not valid_expirations:
        raise ValueError(f'Keine Verfallstermine für "{ticker}" gefunden.')

    # --- Strikes filtern: CSP-typischer Bereich (70–102% des Kurses) ---
    min_strike = current_price * CSP_STRIKE_MIN_FACTOR
    max_strike = current_price * CSP_STRIKE_MAX_FACTOR
    valid_strikes = sorted([s for s in all_strikes_set if min_strike <= s <= max_strike])

    if not valid_strikes:
        raise ValueError(
            f'Keine passenden Strikes für "{ticker}" bei Kurs {current_price:.2f} gefunden.'
        )

    # Strikes auf Maximum begrenzen: gleichmäßig verteilt, höchste (= ATM) bevorzugen
    if len(valid_strikes) > CSP_MAX_STRIKES_PER_EXPIRY:
        step = max(1, len(valid_strikes) // CSP_MAX_STRIKES_PER_EXPIRY)
        reduced = valid_strikes[::step]
        if valid_strikes[-1] not in reduced:
            reduced.append(valid_strikes[-1])
        valid_strikes = sorted(reduced)[-CSP_MAX_STRIKES_PER_EXPIRY:]

    total = len(valid_expirations) * len(valid_strikes)
    status(
        f'{ticker}: {len(valid_expirations)} Verfallstermine × '
        f'{len(valid_strikes)} Strikes = {total} Optionen. Frage Marktdaten ab...'
    )

    # --- Put-Kontrakte erstellen, qualifizieren und Marktdaten abfragen ---
    # tradingClass und multiplier je Expiry aus der zugehörigen Chain lesen.
    opt_contracts = []
    for expiry in valid_expirations:
        c = expiry_to_chain[expiry]
        for strike in valid_strikes:
            opt = Option(ticker, expiry, strike, 'P', best_exchange)
            opt.currency     = currency
            opt.tradingClass = c.tradingClass
            opt.multiplier   = str(c.multiplier)
            opt_contracts.append(opt)

    # In Batches qualifizieren (IB-Limit: ~50 Kontrakte pro Aufruf)
    qualified_opts = []
    for i in range(0, len(opt_contracts), 50):
        batch = opt_contracts[i:i + 50]
        try:
            q = ib.qualifyContracts(*batch)
            qualified_opts.extend(q)
        except Exception:
            pass

    if not qualified_opts:
        raise ValueError(f'Keine qualifizierbaren Put-Optionen für "{ticker}" gefunden.')

    status(f'{ticker}: {len(qualified_opts)} Optionen qualifiziert. Warte auf Marktdaten...')

    opt_ticker_map = {}
    for opt in qualified_opts:
        t = ib.reqMktData(opt, '', False, False)
        opt_ticker_map[opt.conId] = (opt, t)

    time.sleep(MARKET_DATA_WAIT)
    ib.sleep(0)

    # Marktdaten-Abonnements beenden
    for opt, _ in opt_ticker_map.values():
        ib.cancelMktData(opt)

    # --- Ergebnisse aufbereiten ---
    results = []
    for con_id, (opt, t_obj) in opt_ticker_map.items():
        bid = t_obj.bid if t_obj.bid and t_obj.bid > 0 else None
        # Börsenschluss: kein Bid → Fallback auf letzten gehandelten Preis
        is_closing = bid is None
        effective_price = bid if bid is not None else get_price(t_obj)
        if effective_price is None:
            continue  # Kein Preis verfügbar → überspringen

        expiry = getattr(opt, 'lastTradeDateOrContractMonth', '')
        strike = float(getattr(opt, 'strike', 0))
        days   = dte(expiry)

        rr           = (365.0 / days) * (effective_price / strike) if days > 0 and strike > 0 else None
        pct_to_price = (strike / current_price - 1.0) * 100.0 if current_price > 0 else None

        results.append({
            'symbol':        fmt_option_symbol(ticker, expiry, strike, 'P'),
            'strike':        strike,
            'pct_to_price':  pct_to_price,
            'dte':           days,
            'expiry':        expiry,
            'bid':           effective_price,
            'is_closing':    is_closing,
            'restrendite':   rr,
            'exchange':      best_exchange,
        })

    # Sortierung: DTE aufsteigend (= wöchentlich), dann Strike absteigend (höhere zuerst)
    results.sort(key=lambda r: (r['dte'], -r['strike']))

    return {
        'ticker':          ticker,
        'long_name':       long_name,
        'current_price':   current_price,
        'currency':        currency,
        'is_market_closed': is_market_closed,
        'options':         results,
    }


# ---------------------------------------------------------------------------
# TKinter GUI
# ---------------------------------------------------------------------------

class App(tk.Tk):
    """Hauptfenster des IB Ausstiegsrechners."""

    COLUMNS = (
        'symbol', 'bezeichnung', 'typ', 'position', 'kurs',
        'strike', 'dte', 'ablauf', 'kaufpreis', 'optpreis',
        'restrendite', 'waehrung',
    )
    COL_HEADS = {
        'symbol':      'Symbol',
        'bezeichnung': 'Bezeichnung',
        'typ':         'Typ',
        'position':    'Position',
        'kurs':        'Akt. Kurs',
        'strike':      'Strike',
        'dte':         'DTE',
        'ablauf':      'Ablaufdatum',
        'kaufpreis':   'Kauf-/Vkf-Preis',
        'optpreis':    'Akt. Opt-Preis',
        'restrendite': 'Restrendite',
        'waehrung':    'Währung',
    }
    COL_WIDTHS = {
        'symbol': 200, 'bezeichnung': 210, 'typ': 55, 'position': 65,
        'kurs': 90, 'strike': 80, 'dte': 45, 'ablauf': 90,
        'kaufpreis': 110, 'optpreis': 110, 'restrendite': 90, 'waehrung': 65,
    }

    def __init__(self):
        super().__init__()
        self.title(f'IB Ausstiegsrechner  v{APP_VERSION}')
        self.geometry('1400x700')
        self._data = None
        self._ib   = None   # Aktive IB-Verbindung (bleibt nach [Laden] offen)
        self.protocol('WM_DELETE_WINDOW', self._on_close)

        # --- Toolbar ---
        toolbar = tk.Frame(self, bd=1, relief=tk.FLAT)
        toolbar.pack(side=tk.TOP, fill=tk.X, padx=6, pady=6)

        self._btn_laden = tk.Button(
            toolbar, text='Laden', command=self._on_laden, width=10
        )
        self._btn_laden.pack(side=tk.LEFT, padx=(0, 4))

        self._btn_excel = tk.Button(
            toolbar, text='Excel', command=self._on_excel, width=10,
            state=tk.DISABLED
        )
        self._btn_excel.pack(side=tk.LEFT, padx=(0, 4))

        self._btn_csp = tk.Button(
            toolbar, text='CSP Auswahl', command=self._on_csp_auswahl, width=12,
            state=tk.DISABLED
        )
        self._btn_csp.pack(side=tk.LEFT, padx=(0, 10))

        self._status_var = tk.StringVar(value='Bereit. Klicke [Laden] um Daten abzurufen.')
        tk.Label(toolbar, textvariable=self._status_var, anchor='w').pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        # --- Treeview mit Scrollbars ---
        frame = tk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)

        self._tree = ttk.Treeview(
            frame,
            columns=self.COLUMNS,
            show='headings',
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._tree.yview)
        hsb.config(command=self._tree.xview)

        for col in self.COLUMNS:
            anchor = 'w' if col in ('symbol', 'bezeichnung') else 'e'
            self._tree.heading(col, text=self.COL_HEADS[col])
            self._tree.column(
                col, width=self.COL_WIDTHS.get(col, 100),
                minwidth=40, anchor=anchor, stretch=False
            )

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._tree.pack(fill=tk.BOTH, expand=True)

        # --- Farb-Tags konfigurieren ---
        bold = ('Helvetica', 9, 'bold')
        normal = ('Helvetica', 9)
        self._tree.tag_configure('header_cash',
                                 background='#375623', foreground='white', font=bold)
        self._tree.tag_configure('header_csp',
                                 background='#C0504D', foreground='white', font=bold)
        self._tree.tag_configure('row_csp_sum',
                                 background='#E0E0E0', foreground='black', font=bold)
        self._tree.tag_configure('header_stocks',
                                 background='#17375E', foreground='white', font=bold)
        self._tree.tag_configure('header_group_eur',
                                 background='#FFF2CC', foreground='black', font=bold)
        self._tree.tag_configure('header_group_usd',
                                 background='#DEEAF1', foreground='black', font=bold)
        self._tree.tag_configure('row_cash',   background='#EBF1DE', font=normal)
        self._tree.tag_configure('row_csp_0',      background='#C6EFCE',  font=normal)  # OTM: hellgrün
        self._tree.tag_configure('row_csp_1',      background='#A9D18E',  font=normal)  # OTM: mittelgrün
        self._tree.tag_configure('row_csp_0_itm',  background='#FFB3B3',  font=normal)  # ITM: hellrot
        self._tree.tag_configure('row_csp_1_itm',  background='#FF9999',  font=normal)  # ITM: mittelrot
        self._tree.tag_configure('row_stock',        background='#DCE6F1',  font=bold)
        self._tree.tag_configure('row_stock_profit', background='#C6EFCE',  font=bold)
        self._tree.tag_configure('row_stock_loss',   background='#FFC7CE',  font=bold)
        self._tree.tag_configure('row_call',          background='#EBF1DE',  font=normal)  # Short/Covered Call OTM
        self._tree.tag_configure('row_call_itm',      background='#FFB3B3',  font=normal)  # Short/Covered Call ITM
        self._tree.tag_configure('row_call_long',      background='#C6EFCE',  font=normal)  # Long Call: Gewinn
        self._tree.tag_configure('row_call_long_itm', background='#A9D18E',  font=normal)  # Long Call: Gewinn + ITM
        self._tree.tag_configure('row_call_long_loss', background='#FFB3B3', font=normal)  # Long Call: Verlust
        self._tree.tag_configure('row_sep',    background='#D9D9D9',  font=normal)

    # ------------------------------------------------------------------
    # Toolbar-Aktionen
    # ------------------------------------------------------------------

    def _on_laden(self):
        self._btn_laden.config(state=tk.DISABLED)
        self._btn_excel.config(state=tk.DISABLED)
        self._status_var.set('Verbinde...')
        t = threading.Thread(target=self._load_in_thread, daemon=True)
        t.start()

    def _on_excel(self):
        if self._data is None:
            messagebox.showwarning('Kein Daten', 'Bitte zuerst Daten laden.')
            return
        try:
            write_excel(self._data, OUTPUT_FILE)
            messagebox.showinfo('Excel Export', f'Gespeichert: {OUTPUT_FILE}')
        except Exception as e:
            messagebox.showerror('Fehler', f'Excel-Export fehlgeschlagen:\n{e}')

    # ------------------------------------------------------------------
    # Hintergrund-Thread: Daten laden
    # ------------------------------------------------------------------

    def _load_in_thread(self):
        # Jeder Thread benötigt eine eigene asyncio-Event-Loop
        asyncio.set_event_loop(asyncio.new_event_loop())

        # Bestehende Verbindung trennen – sonst lehnt IB die gleiche CLIENT_ID ab
        if self._ib is not None:
            try:
                self._ib.disconnect()
            except Exception:
                pass
            self._ib = None

        ib = IB()
        SUPPRESS_CODES = _BENIGN_LOG_CODES | {2104, 2106, 2107, 2108, 2158}

        def error_handler(req_id, code, msg, contract=None):
            if code in SUPPRESS_CODES:
                return
            self.after(0, lambda m=msg, c=code: self._status_var.set(
                f'TWS Fehler {c}: {m}'
            ))

        ib.errorEvent += error_handler

        try:
            self.after(0, lambda: self._status_var.set(
                f'Verbinde mit TWS auf {TWS_HOST}:{TWS_PORT}...'
            ))
            ib.connect(TWS_HOST, TWS_PORT, clientId=CLIENT_ID, timeout=10, readonly=True)
        except Exception as e:
            self.after(0, lambda err=e: self._on_load_error(
                f'Verbindung zu TWS fehlgeschlagen:\n{err}\n\n'
                'Bitte sicherstellen, dass TWS läuft und die API aktiviert ist.'
            ))
            return

        try:
            def status_cb(msg: str):
                self.after(0, lambda m=msg: self._status_var.set(m))

            data = collect_data(ib, status_callback=status_cb)
            if data is None:
                ib.disconnect()
                self.after(0, lambda: self._on_load_error('Keine offenen Positionen gefunden.'))
            else:
                # Verbindung offen lassen – wird für CSP-Auswahl und Fenster-Close benötigt
                self._ib = ib
                self.after(0, lambda d=data: self._update_table(d))
        except Exception as e:
            ib.disconnect()
            self.after(0, lambda err=e: self._on_load_error(f'Fehler beim Laden:\n{err}'))

    def _on_load_error(self, msg: str):
        self._status_var.set('Fehler.')
        self._btn_laden.config(state=tk.NORMAL)
        messagebox.showerror('Fehler', msg)

    def _on_close(self):
        """Fenster schließen: IB-Verbindung sauber trennen."""
        ib = getattr(self, '_ib', None)
        if ib is not None:
            try:
                ib.disconnect()
            except Exception:
                pass
        self.destroy()

    def _on_csp_auswahl(self):
        """Öffnet den CSP-Auswahl-Dialog."""
        CSPAuswahlDialog(self)

    # ------------------------------------------------------------------
    # Tabelle befüllen
    # ------------------------------------------------------------------

    def _update_table(self, data: dict):
        self._data = data
        tree = self._tree

        # Alle vorhandenen Zeilen löschen
        tree.delete(*tree.get_children())

        def ins(values: tuple, tag: str):
            """Fügt eine Zeile ein; füllt fehlende Spalten mit leerem String auf."""
            padded = list(values) + [''] * (len(self.COLUMNS) - len(values))
            tree.insert('', 'end', values=padded[:len(self.COLUMNS)], tags=(tag,))

        def fmt_num(val, fmt='.2f') -> str:
            if val is None:
                return 'n/v'
            return format(val, fmt)

        def fmt_pct(val) -> str:
            if val is None:
                return '-'
            return f'{val:.2%}'

        cash_balance      = data['cash_balance']
        csp_margin        = data['csp_margin']
        stock_capital     = data['stock_capital']
        long_call_capital = data['long_call_capital']
        free_cash         = data['free_cash']
        csp_rows     = data['csp_rows']
        stock_map    = data['stock_map']
        call_map     = data['call_map']
        syms_eur     = data['syms_eur']
        syms_usd     = data['syms_usd']
        eurusd_price = data['eurusd_price']
        now_str      = data['timestamp']

        # ===========================================================
        # Abschnitt: Kontoguthaben
        # ===========================================================
        ins(('Kontoguthaben',), 'header_cash')

        # EUR-Gruppe
        ins(('EUR-Guthaben',), 'header_group_eur')
        ins(('Barmittel',            f"{cash_balance['EUR']:>16,.2f} EUR"), 'row_cash')
        ins(('  CSP-Kapital',        f"{csp_margin['EUR']:>16,.2f} EUR"), 'row_cash')
        ins(('  Aktien-Kapital',     f"{stock_capital['EUR']:>16,.2f} EUR"), 'row_cash')
        ins(('  Long Call-Kapital',  f"{long_call_capital['EUR']:>16,.2f} EUR"), 'row_cash')
        ins(('Freier Cash',          f"{free_cash['EUR']:>16,.2f} EUR"), 'row_cash')

        # Trennlinie zwischen EUR und USD
        ins((), 'row_sep')

        # USD-Gruppe
        ins(('USD-Guthaben',), 'header_group_usd')
        ins(('Barmittel',            f"{cash_balance['USD']:>16,.2f} USD"), 'row_cash')
        ins(('  CSP-Kapital',        f"{csp_margin['USD']:>16,.2f} USD"), 'row_cash')
        ins(('  Aktien-Kapital',     f"{stock_capital['USD']:>16,.2f} USD"), 'row_cash')
        ins(('  Long Call-Kapital',  f"{long_call_capital['USD']:>16,.2f} USD"), 'row_cash')
        ins(('Freier Cash',          f"{free_cash['USD']:>16,.2f} USD"), 'row_cash')

        # Statuszeile: Stand und EUR/USD
        ins((f'Stand: {now_str}   |   EUR/USD: {eurusd_price:.4f}',), 'row_cash')

        # Leerzeile
        ins((), 'row_sep')

        # ===========================================================
        # Abschnitt: Cash Secured Puts (CSPs)
        # ===========================================================
        ins(('Cash Secured Puts (CSPs)',), 'header_csp')

        last_currency = None
        csp_color_idx = 0  # eigener Zähler für alternierende Farben (unabhängig vom Trenner)
        for row in csp_rows:
            # Währungswechsel: Summe der Vorgängergruppe + Unterüberschrift einfügen
            if row['currency'] != last_currency:
                if last_currency is not None:
                    ins(('Σ Gebundenes Kapital',
                         f"{csp_margin[last_currency]:>16,.2f} {last_currency}"),
                        'row_csp_sum')
                    ins((), 'row_sep')
                group_tag = 'header_group_usd' if row['currency'] == 'USD' else 'header_group_eur'
                ins((f"{row['currency']}-Positionen",), group_tag)
                last_currency = row['currency']
                csp_color_idx = 0
            ul_price    = row['underlying_price']
            cur_opt     = row['current_price']
            restrendite = calc_restrendite(
                row['premium'], row['strike'], row['dte'], cur_opt
            )
            # Farbe nach Gewinn/Verlust: rot wenn aktueller Optionspreis >= erhaltene Prämie
            is_loss = cur_opt is not None and cur_opt >= row['premium']
            if is_loss:
                tag = 'row_csp_0_itm' if csp_color_idx % 2 == 0 else 'row_csp_1_itm'
            else:
                tag = 'row_csp_0' if csp_color_idx % 2 == 0 else 'row_csp_1'
            csp_color_idx += 1
            ins((
                row['display_symbol'],
                row['bezeichnung'],
                'CSP',
                str(int(row['position'])),
                fmt_num(ul_price),
                fmt_num(row['strike']),
                str(row['dte']),
                fmt_date(row['expiry']),
                fmt_num(row['premium']),
                fmt_num(cur_opt),
                fmt_pct(restrendite),
                row['currency'],
            ), tag)

        # Summe der letzten Währungsgruppe
        if last_currency is not None:
            ins(('Σ Gebundenes Kapital',
                 f"{csp_margin[last_currency]:>16,.2f} {last_currency}"),
                'row_csp_sum')

        # Leerzeile
        ins((), 'row_sep')

        # ===========================================================
        # Abschnitt: Aktien & Covered Calls
        # ===========================================================
        ins(('Aktien & Covered Calls',), 'header_stocks')

        def write_group(sym_list: list, group_label: str, group_tag: str):
            if not sym_list:
                return
            ins((group_label,), group_tag)

            for sym_idx, sym in enumerate(sym_list):

                # Aktienzeile – Farbe je nach Gewinn/Verlust (aktueller Kurs vs. Kaufpreis)
                if sym in stock_map:
                    s = stock_map[sym]
                    cp, ac = s['current_price'], s['avg_cost']
                    if cp and ac and ac > 0:
                        stock_tag = 'row_stock_profit' if cp > ac else 'row_stock_loss'
                    else:
                        stock_tag = 'row_stock'
                    ins((
                        s['display_symbol'],
                        s['bezeichnung'],
                        'Aktie',
                        str(int(s['position'])),
                        fmt_num(cp),
                        '-', '-', '-',
                        fmt_num(ac),
                        '-', '-',
                        s['currency'],
                    ), stock_tag)

                # Call-Zeilen
                if sym in call_map:
                    ul_price_for_call = stock_map.get(sym, {}).get('current_price')
                    for call_row in call_map[sym]:
                        cur_opt  = call_row['current_price']
                        is_long  = call_row.get('is_long', False)
                        is_itm   = (ul_price_for_call is not None
                                    and ul_price_for_call > call_row['strike'])

                        if is_long:
                            # Gekaufter Call: nur anzeigen wenn Optionspreis > Einkaufspreis
                            # (analog zu Short-Optionen, die nur bei Gewinn anzeigen)
                            if (cur_opt is not None and call_row['premium'] > 0
                                    and cur_opt > call_row['premium']):
                                pnl_pct = (cur_opt - call_row['premium']) / call_row['premium']
                                rr_str  = fmt_pct(pnl_pct)
                            else:
                                rr_str = '-'
                            is_profit = (cur_opt is not None
                                         and cur_opt > call_row['premium'])
                            if not is_profit:
                                call_tag = 'row_call_long_loss'
                            elif is_itm:
                                call_tag = 'row_call_long_itm'
                            else:
                                call_tag = 'row_call_long'
                            typ_str  = 'Call (L)'
                        else:
                            # Verkaufter Call (Covered Call): Restrendite wie gehabt
                            restrendite = calc_restrendite(
                                call_row['premium'], call_row['strike'],
                                call_row['dte'], cur_opt
                            )
                            rr_str   = fmt_pct(restrendite)
                            call_tag = 'row_call_itm' if is_itm else 'row_call'
                            typ_str  = 'Call'

                        ins((
                            call_row['display_symbol'],
                            call_row['bezeichnung'],
                            typ_str,
                            str(int(call_row['position'])),
                            fmt_num(ul_price_for_call),
                            fmt_num(call_row['strike']),
                            str(call_row['dte']),
                            fmt_date(call_row['expiry']),
                            fmt_num(call_row['premium']),
                            fmt_num(cur_opt),
                            rr_str,
                            call_row['currency'],
                        ), call_tag)

                # Trennzeile zwischen Symbolen
                if sym_idx < len(sym_list) - 1:
                    ins((), 'row_sep')

        write_group(syms_eur, 'EUR-Positionen', 'header_group_eur')
        write_group(syms_usd, 'USD-Positionen', 'header_group_usd')

        # Status aktualisieren
        n_csps  = len(csp_rows)
        n_stk   = len(stock_map)
        n_calls = sum(len(v) for v in call_map.values())
        self._status_var.set(
            f'Geladen: {n_csps} CSPs, {n_stk} Aktien, {n_calls} Calls  |  Stand: {now_str}'
        )
        self._btn_laden.config(state=tk.NORMAL)
        self._btn_excel.config(state=tk.NORMAL)
        self._btn_csp.config(state=tk.NORMAL)


# ---------------------------------------------------------------------------
# CSP-Auswahl-Dialog
# ---------------------------------------------------------------------------

class CSPAuswahlDialog(tk.Toplevel):
    """Dialogfenster zur Suche von CSP-Kandidaten für einen beliebigen Ticker.

    Der Nutzer gibt ein Ticker-Symbol ein. Das Programm lädt dann alle
    Short-Put-Optionen mit DTE ≤ 60 Tage und Strikes im Bereich 70–102%
    des aktuellen Kurses und zeigt die erwartete Restrendite an.

    Die Suche läuft in einem eigenen Hintergrund-Thread mit einer separaten
    IB-Verbindung (CSP_CLIENT_ID), damit die Hauptanwendung nicht blockiert.
    """

    COLUMNS = ('symbol', 'strike', 'pct_kurs', 'dte', 'ablauf', 'bid', 'restrendite', 'boerse')
    COL_HEADS = {
        'symbol':      'Symbol',
        'strike':      'Strike',
        'pct_kurs':    '% zum Kurs',
        'dte':         'DTE',
        'ablauf':      'Ablaufdatum',
        'bid':         'Bid (Prämie)',
        'restrendite': 'Restrendite p.a.',
        'boerse':      'Börse',
    }
    COL_WIDTHS = {
        'symbol': 230, 'strike': 80, 'pct_kurs': 90, 'dte': 50,
        'ablauf': 100, 'bid': 100, 'restrendite': 120, 'boerse': 80,
    }

    def __init__(self, parent_app: 'App'):
        super().__init__(parent_app)
        self._app = parent_app
        self.title('CSP Auswahl')
        self.geometry('1000x560')
        self.transient(parent_app)

        # --- Info-Header: Ticker | Firmenname | Aktueller Kurs ---
        info_frame = tk.Frame(self, bg='#17375E', pady=5)
        info_frame.pack(fill=tk.X)
        self._info_var = tk.StringVar(value='')
        tk.Label(
            info_frame, textvariable=self._info_var,
            bg='#17375E', fg='white',
            font=('Helvetica', 11, 'bold'), anchor='w', padx=10,
        ).pack(fill=tk.X)

        # --- Eingabezeile ---
        input_frame = tk.Frame(self)
        input_frame.pack(fill=tk.X, padx=8, pady=8)

        tk.Label(input_frame, text='Ticker:').pack(side=tk.LEFT)
        self._ticker_var = tk.StringVar()
        entry = tk.Entry(input_frame, textvariable=self._ticker_var, width=12)
        entry.pack(side=tk.LEFT, padx=4)
        entry.bind('<Return>', lambda _e: self._on_suchen())

        self._btn_suchen = tk.Button(
            input_frame, text='Suchen', command=self._on_suchen, width=8
        )
        self._btn_suchen.pack(side=tk.LEFT, padx=(0, 10))

        self._status_var = tk.StringVar(value='Ticker eingeben und [Suchen] klicken.')
        tk.Label(input_frame, textvariable=self._status_var, anchor='w').pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        # --- Treeview ---
        frame = tk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        self._tree = ttk.Treeview(
            frame,
            columns=self.COLUMNS,
            show='headings',
            yscrollcommand=vsb.set,
        )
        vsb.config(command=self._tree.yview)

        for col in self.COLUMNS:
            anchor = 'w' if col == 'symbol' else 'e'
            self._tree.heading(col, text=self.COL_HEADS[col])
            self._tree.column(
                col, width=self.COL_WIDTHS.get(col, 100),
                minwidth=40, anchor=anchor, stretch=False,
            )

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._tree.pack(fill=tk.BOTH, expand=True)

        # Farb-Tags: grün = gute Rendite (≥5%), gelb = niedrig, rosa = keine Rendite
        bold = ('Helvetica', 9, 'bold')
        self._tree.tag_configure('good', background='#EBF1DE', font=bold)
        self._tree.tag_configure('low',  background='#FFF2CC')
        self._tree.tag_configure('none', background='#F2DCDB')

        entry.focus_set()

    # ------------------------------------------------------------------
    # Suchen-Aktion
    # ------------------------------------------------------------------

    def _on_suchen(self):
        ticker = self._ticker_var.get().strip().upper()
        if not ticker:
            return
        self._btn_suchen.config(state=tk.DISABLED)
        self._status_var.set(f'Suche {ticker}...')
        self._tree.delete(*self._tree.get_children())
        t = threading.Thread(target=self._search_thread, args=(ticker,), daemon=True)
        t.start()

    def _search_thread(self, ticker: str):
        """Hintergrund-Thread: Eigene IB-Verbindung aufbauen und CSPs laden."""
        asyncio.set_event_loop(asyncio.new_event_loop())

        ib = IB()
        SUPPRESS_CODES = _BENIGN_LOG_CODES | {2104, 2106, 2107, 2108, 2158}

        def error_handler(req_id, code, msg, contract=None):
            if code in SUPPRESS_CODES:
                return
            self.after(0, lambda m=msg, c=code: self._status_var.set(
                f'TWS Fehler {c}: {m}'
            ))

        ib.errorEvent += error_handler

        try:
            self.after(0, lambda: self._status_var.set(
                f'Verbinde mit TWS (clientId={CSP_CLIENT_ID})...'
            ))
            ib.connect(TWS_HOST, TWS_PORT, clientId=CSP_CLIENT_ID, timeout=10, readonly=True)
        except Exception as e:
            self.after(0, lambda err=e: self._on_error(
                f'Verbindung zu TWS fehlgeschlagen:\n{err}'
            ))
            return

        try:
            def status_cb(msg: str):
                self.after(0, lambda m=msg: self._status_var.set(m))

            result_data = fetch_csp_candidates(
                ib, ticker, self._app._data, status_callback=status_cb
            )
            self.after(0, lambda d=result_data: self._show_results(d))
        except ValueError as e:
            self.after(0, lambda err=e: self._on_error(str(err)))
        except Exception as e:
            self.after(0, lambda err=e: self._on_error(f'Unerwarteter Fehler:\n{err}'))
        finally:
            ib.disconnect()

    def _on_error(self, msg: str):
        self._status_var.set('Fehler.')
        self._btn_suchen.config(state=tk.NORMAL)
        messagebox.showerror('Fehler', msg, parent=self)

    # ------------------------------------------------------------------
    # Ergebnisse anzeigen
    # ------------------------------------------------------------------

    def _show_results(self, data: dict):
        ticker            = data['ticker']
        long_name         = data['long_name']
        current_price     = data['current_price']
        currency          = data['currency']
        is_market_closed  = data.get('is_market_closed', False)
        options           = data['options']

        # Info-Header aktualisieren; bei Börsenschluss "(Schlusskurs)" anzeigen
        kurs_label = f'{current_price:.2f} {currency}'
        if is_market_closed:
            kurs_label += '  (Schlusskurs)'
        self._info_var.set(f'{ticker}   |   {long_name}   |   Kurs: {kurs_label}')

        self._tree.delete(*self._tree.get_children())

        if not options:
            self._status_var.set(
                f'Keine CSP-Optionen mit Preis für "{ticker}" gefunden.'
            )
            self._btn_suchen.config(state=tk.NORMAL)
            return

        for row in options:
            rr  = row['restrendite']
            pct = row['pct_to_price']
            if rr is None or row.get('is_closing'):
                tag = 'none'
            elif rr < 0.05:
                tag = 'low'
            else:
                tag = 'good'

            # Bid-Spalte: bei Schlusskurs mit "(S)" markieren
            bid_str = f"{row['bid']:.2f}"
            if row.get('is_closing'):
                bid_str += ' (S)'

            self._tree.insert('', 'end', values=(
                row['symbol'],
                f"{row['strike']:.2f}",
                f"{pct:+.1f}%" if pct is not None else '-',
                str(row['dte']),
                fmt_date(row['expiry']),
                bid_str,
                f"{rr:.2%}" if rr is not None else '-',
                row['exchange'],
            ), tags=(tag,))

        closing_count = sum(1 for r in options if r.get('is_closing'))
        status = f'{len(options)} CSP-Optionen für "{ticker}" gefunden  (grün ≥ 5% p.a., gelb < 5% p.a.)'
        if closing_count:
            status += f'  |  {closing_count}× (S) = Schlusskurs, kein live Bid'
        self._status_var.set(status)
        self._btn_suchen.config(state=tk.NORMAL)


# ---------------------------------------------------------------------------
# Einstiegspunkt
# ---------------------------------------------------------------------------

def main():
    """Startet die GUI-Anwendung."""
    app = App()
    app.mainloop()


if __name__ == '__main__':
    main()
